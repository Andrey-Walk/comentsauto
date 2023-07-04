import io
from django.http import HttpResponse
import openpyxl
from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework import viewsets
from .models import Countrys, Directors, Cars, Comments
from .serializers import DirectorsSerializer, CountrysSerializer, CarsSerializer, CommentsSerializer
from .permissions import LookPermission, TokenPermission


def index(request):
    news = Cars.objects.all()
    return render(request, 'main/index.html', {'main': news})

class DirectorsModelViewSet(viewsets.ModelViewSet):
    queryset = Directors.objects.all()
    serializer_class = DirectorsSerializer

class CountrysModelViewSet(viewsets.ModelViewSet):
    queryset = Countrys.objects
    serializer_class = CountrysSerializer
    permission_classes = [LookPermission|TokenPermission]

class CarsModelViewSet(viewsets.ModelViewSet):
    queryset = Cars.objects
    serializer_class = CarsSerializer
    permission_classes = [LookPermission|TokenPermission]

class CommentsModelViewSet(viewsets.ModelViewSet):
    queryset = Comments.objects
    serializer_class = CommentsSerializer

class CountrysAPIView(APIView):
    def get(self, request):
        format_file = request.query_params.get('param')
        file_extension = format_file
        queryset = Countrys.objects.all()
        data = CountrysSerializer(queryset, many=True).data
        book = openpyxl.Workbook()
        sheet_1 = book.active
        sheet_1.title = 'Страны'
        for item in data:
            sheet_1.append([item['country']])
        file_stream = io.BytesIO()
        book.save(file_stream)
        file_stream.seek(0)
        response = HttpResponse(file_stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="countrys.{file_extension}"'

        return response

class DirectorsAPIView(APIView):
    def get(self, request):
        format_file = request.query_params.get('param')
        file_extension = format_file
        queryset = Directors.objects.all()
        data = DirectorsSerializer(queryset, many=True).data
        book = openpyxl.Workbook()
        sheet_1 = book.active
        sheet_1.title = 'Производители'
        for item in data:
            sheet_1.append([item['director'], item['country']])
        file_stream = io.BytesIO()
        book.save(file_stream)
        file_stream.seek(0)
        response = HttpResponse(file_stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="directors.{file_extension}"'

        return response

class CarsAPIView(APIView):
    def get(self, request):
        format_file = request.query_params.get('param')
        file_extension = format_file
        queryset = Cars.objects.all()
        data = CarsSerializer(queryset, many=True).data
        book = openpyxl.Workbook()
        sheet_1 = book.active
        sheet_1.title = 'Автомобили'
        for item in data:
            sheet_1.append([item['car'], item['director'], item['datastart'], item['dataend']])
        file_stream = io.BytesIO()
        book.save(file_stream)
        file_stream.seek(0)
        response = HttpResponse(file_stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="cars.{file_extension}"'

        return response

class CommentsAPIView(APIView):
    def get(self, request):
        format_file = request.query_params.get('param')
        file_extension = format_file
        queryset = Comments.objects.all()
        data = CommentsSerializer(queryset, many=True).data
        book = openpyxl.Workbook()
        sheet_1 = book.active
        sheet_1.title = 'Комментарии'
        for item in data:
            sheet_1.append([item['avtor'], item['avtocar'], item['datacreate'], item['comment']])
        file_stream = io.BytesIO()
        book.save(file_stream)
        file_stream.seek(0)
        response = HttpResponse(file_stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="сomments.{file_extension}"'

        return response

class AllAPIView(APIView):
    def get(self, request):
        format_file = request.query_params.get('param')
        file_extension = format_file
        queryset = Countrys.objects.all()
        data = CountrysSerializer(queryset, many=True).data
        book = openpyxl.Workbook()
        sheet_1 = book.create_sheet(title='Страны', index=0)
        for item in data:
            sheet_1.append([item['country']])

        queryset = Directors.objects.all()
        data = DirectorsSerializer(queryset, many = True).data
        sheet_2 = book.create_sheet(title='Производители', index=1)
        for item in data:
            sheet_2.append([item['director'], item['country']])

        queryset = Cars.objects.all()
        data = CarsSerializer(queryset, many = True).data
        sheet_3 = book.create_sheet(title='Автомобили', index=2)
        for item in data:
            sheet_3.append([item['car'], item['director'], item['datastart'], item['dataend']])

        queryset = Comments.objects.all()
        data = CommentsSerializer(queryset, many = True).data
        sheet_4 = book.create_sheet(title='Комментарии', index=3)
        for item in data:
            sheet_4.append([item['avtor'], item['avtocar'], item['datacreate'], item['comment']])

        file_stream = io.BytesIO()
        book.save(file_stream)
        file_stream.seek(0)
        response = HttpResponse(file_stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="all.{file_extension}"'

        return response